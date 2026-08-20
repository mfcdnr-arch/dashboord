"""Экспорт данных страницы в .xlsx (вынесено из _widgetdata.py)."""
from __future__ import annotations

import re
from datetime import date

from ._base import DashboardError
from ._rls import _can_view, visible_widget_ids
from ._sheetnames import LIMIT, clean_title, short_cores
from ._widgetdata import compute_widget_data
from ._widgetsources import _page_org

_ISO_DATE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _as_date(v):
    """ISO-строку периода превращаем в НАСТОЯЩУЮ дату Excel.

    Раньше в файл уезжал текст «2026-07-22»: на экране ДД.ММ.ГГГГ, в файле
    что-то другое, а сортировка и фильтр по такому столбцу работают как по
    строке. Настоящая дата решает всё сразу — и вид (формат ячейки), и
    сортировку. «2026-07» (месяц, а не день) НЕ трогаем: превращать его в дату
    значило бы выдумать день.
    """
    if isinstance(v, str) and _ISO_DATE.match(v):
        try:
            return date.fromisoformat(v)
        except ValueError:
            return v
    return v


def _ru_date(v) -> str:
    """Дата внутри текстовой подписи — как на экране: ДД.ММ.ГГГГ."""
    d = _as_date(v)
    return d.strftime("%d.%m.%Y") if isinstance(d, date) else ("" if v is None else str(v))


# Строгий ДД.ММ.ГГГГ (решение заказчика 19.08) — но точки в коде ОБЯЗАТЕЛЬНО
# экранированы. Разница не косметическая, она проверена экспериментом: файл с
# одной и той же датой в четырёх записях формата прочитан сторонним движком,
# на котором работают веб-просмотрщики, и «dd.mm.yyyy» он отдал голым числом
# «46225» (та же беда, что заказчик видел как «DD.07.YYYY»), а «dd"."mm"."yyyy»
# — датой «22.07.2026». Причина: неэкранированная точка читается как
# десятичный разделитель и ломает разбор всего шаблона. Встроенный формат
# (numFmtId=14) тоже рисуется везде, но ПО НАСТРОЙКАМ СИСТЕМЫ, а заказчику
# нужен русский вид независимо от программы — поэтому свой код.
DATE_FMT = 'dd"."mm"."yyyy'


def _row(ws, values: list) -> None:
    """Строка данных: даты кладём датами и подписываем форматом ДД.ММ.ГГГГ."""
    ws.append([_as_date(v) for v in values])
    for c in ws[ws.max_row]:
        if isinstance(c.value, date):
            c.number_format = DATE_FMT


def _autofit(ws) -> None:
    """Ширина столбцов по содержимому.

    По умолчанию openpyxl оставляет 8 знаков, и подпись «За весь период
    (22.07.2026 → 19.08.2026)» в файле обрезалась до «За весь пер» — то есть
    строка итога переставала читаться. Потолок нужен, иначе имя показателя
    госформы растянет столбец на пол-экрана.
    """
    from openpyxl.utils import get_column_letter

    for i, col in enumerate(ws.iter_cols(), 1):
        best = 0
        for c in col:
            if c.value is None:
                continue
            n = 10 if isinstance(c.value, date) else len(str(c.value))
            best = max(best, n)
        if best:
            ws.column_dimensions[get_column_letter(i)].width = min(max(best + 2, 10), 60)


def _pct(v: float | None) -> str:
    """Процент в ячейку итогов: «+4,28 %» либо пусто, если посчитать не от чего."""
    return "" if v is None else f"{v:+.2f} %".replace(".", ",")


async def export_page_xlsx(conn, org_id, user: dict, page_id: str) -> bytes:
    """Экспорт данных всех виджетов страницы в .xlsx (openpyxl).
    KPI/план-факт — на лист «Сводка», датасетные виджеты — по листу на виджет.
    Аннотации (text/image) пропускаются. RLS: проверяется доступ к дашборду."""
    import io

    from openpyxl import Workbook

    p = await _page_org(conn, org_id, page_id)
    if p is None:
        raise DashboardError("Страница не найдена")
    if not await _can_view(conn, org_id, user, str(p["dashboard_id"])):
        raise DashboardError("Страница не найдена")

    allowed = await visible_widget_ids(conn, org_id, user, str(p["dashboard_id"]))
    rows = await conn.fetch(
        "select id, name, widget_type from widgets where page_id=$1::uuid order by position_y, position_x", page_id)
    if allowed is not None:
        rows = [w for w in rows if str(w["id"]) in allowed]

    wb = Workbook()
    summary = wb.active
    summary.title = "Сводка"
    summary.append(["Виджет", "Тип", "Показатель", "Значение"])
    has_summary = False

    # Имя листа Excel — не больше 31 знака, а имя показателя госформы втрое
    # длиннее. Обрезка слева давала тринадцать листов «Динамика  Количество
    # поль 2 / 3 / 4» — понять, какой лист про что, было нельзя. Теперь имя
    # собирается из НОМЕРА и различающей части (см. _sheetnames), а полные
    # имена лежат на листе «Содержание».
    cand_rows = [w for w in rows if w["widget_type"] not in ("text", "image", "kpi", "gauge", "plan_fact")]
    cores = dict(zip([str(w["id"]) for w in cand_rows],
                     short_cores([w["name"] for w in cand_rows], LIMIT - 3), strict=True))
    made: list[tuple[str, str]] = []
    used: set = {"сводка", "содержание"}
    seq = 0
    def sheet_name(wid: str, base: str) -> str:
        nonlocal seq
        seq += 1
        cand = f"{seq:02d} {cores.get(wid) or clean_title(base)}"[:LIMIT]
        stem, i = cand, 2
        while cand.lower() in used:
            cand, i = f"{stem[: LIMIT - 2]} {i}", i + 1
        used.add(cand.lower())
        made.append((cand, base))
        return cand

    for w in rows:
        wid, t, name = str(w["id"]), w["widget_type"], w["name"]
        if t in ("text", "image"):
            continue
        try:
            data = await compute_widget_data(conn, org_id, wid, skip_acl=True)
        except DashboardError:
            continue
        if t == "kpi":
            summary.append([name, "KPI", "значение", data.get("value")]); has_summary = True
        elif t == "gauge":
            summary.append([name, "Спидометр", "значение", data.get("value")]); has_summary = True
        elif t == "plan_fact":
            summary.append([name, "План-факт", "план", data.get("plan")])
            summary.append([name, "План-факт", "факт", data.get("fact")])
            summary.append([name, "План-факт", "выполнение, %", data.get("pct")]); has_summary = True
        elif t == "table":
            ws = wb.create_sheet(sheet_name(wid, name))
            cols = list(data.get("columns", []))
            # Заголовки — ИМЕНА показателей, а не коды полей: на экране таблица
            # подписана именами (column_titles, 2026-08-08), и файл обязан
            # совпадать с тем, что человек видел. Код остаётся ключом строки.
            titles = data.get("column_titles") or {}
            ws.append(["Строка"] + [titles.get(c, c) for c in cols])
            for r in data.get("rows", []):
                ws.append([r.get("row")] + [r.get(c) for c in cols])
        elif t in ("bar", "line", "pie"):
            ws = wb.create_sheet(sheet_name(wid, name))
            ws.append(["Категория", "Значение"])
            for c, v in zip(data.get("categories", []), data.get("values", []), strict=False):
                ws.append([c, v])
        elif t == "dynamics":
            ws = wb.create_sheet(sheet_name(wid, name))
            anomaly_idx = {a["index"] for a in data.get("anomalies", [])}
            ws.append(["Период", "Значение", "Аномалия"])
            for i, (pr, v) in enumerate(zip(data.get("periods", []), data.get("values", []), strict=False)):
                _row(ws, [pr, v, "⚠" if i in anomaly_idx else ""])
            # Те же итоги, что видны под графиком — иначе в выгрузке пришлось бы
            # считать их заново вручную, и цифры разошлись бы с экраном.
            if data.get("total_change") is not None:
                ws.append([])
                ws.append([f"За весь период ({_ru_date(data.get('first_period'))} → "
                           f"{_ru_date(data.get('last_period'))})",
                           data.get("total_change"), _pct(data.get("total_change_pct"))])
            if data.get("change") is not None:
                ws.append([f"К пред. периоду ({_ru_date(data.get('change_from_period'))} → "
                           f"{_ru_date(data.get('change_to_period'))})",
                           data.get("change"), _pct(data.get("change_pct"))])
        elif t == "yoy":
            ws = wb.create_sheet(sheet_name(wid, name))
            py, cy = data.get("previous_year"), data.get("current_year")
            ws.append(["Месяц", str(py) if py else "пред. год", str(cy)])
            for mn, pv, cv in zip(data.get("months", []), data.get("previous", []), data.get("current", []), strict=False):
                ws.append([mn, pv, cv])
        elif t in ("compare", "cross_dataset_compare"):
            ws = wb.create_sheet(sheet_name(wid, name))
            series = data.get("series", [])
            cats = data.get("categories", [])
            ws.append(["Категория"] + [s.get("name") for s in series])
            for i, c in enumerate(cats):
                _row(ws, [c] + [(s.get("data") or [])[i] if i < len(s.get("data", [])) else None for s in series])
        elif t == "heatmap":
            ws = wb.create_sheet(sheet_name(wid, name))
            cols = list(data.get("columns", []))
            rws = list(data.get("rows", []))
            grid = [[None] * len(cols) for _ in rws]
            for ci, ri, v in data.get("cells", []):
                if ri < len(rws) and ci < len(cols):
                    grid[ri][ci] = v
            ws.append(["Строка"] + cols)
            for i, rname in enumerate(rws):
                ws.append([rname] + grid[i])
        elif t == "pivot":
            ws = wb.create_sheet(sheet_name(wid, name))
            cols = list(data.get("columns", []))
            ws.append(["Строка"] + cols + ["Итого"])
            for r in data.get("rows", []):
                ws.append([r.get("row")] + list(r.get("values", [])) + [r.get("total")])
            ws.append(["Итого"] + list(data.get("col_totals", [])) + [data.get("grand_total")])
        elif t == "waterfall":
            ws = wb.create_sheet(sheet_name(wid, name))
            ws.append(["Категория", "Значение"])
            for c, v in zip(data.get("categories", []), data.get("values", []), strict=False):
                ws.append([c, v])
            ws.append([data.get("total_label", "Итого"), sum(v for v in data.get("values", []) if v is not None)])
        elif t == "objects_compare":
            ws = wb.create_sheet(sheet_name(wid, name))
            ws.append(["Подразделение", "Значение"])
            for c, v in zip(data.get("categories", []), data.get("values", []), strict=False):
                ws.append([c, v])

    if not has_summary and len(wb.sheetnames) > 1:
        wb.remove(summary)  # нет KPI/план-факта, но есть датасетные листы — убираем пустую сводку

    # «Содержание» — единственный способ связать короткое имя листа с полным
    # именем виджета: в 31 знак имя госформы не помещается ни при каком
    # сокращении. Ставим первым листом, чтобы файл открывался на нём.
    if made:
        toc = wb.create_sheet("Содержание", 0)
        toc.append(["Лист", "Виджет"])
        for title, full in made:
            toc.append([title, full])
        toc.column_dimensions["A"].width = 34
        toc.column_dimensions["B"].width = 90

    for ws in wb.worksheets:
        _autofit(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
