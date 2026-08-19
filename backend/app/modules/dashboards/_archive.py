"""Архив дашбордов: слепки данных, месячные папки, темы, избирательный доступ,
ежемесячная автоархивация.

Ключевой принцип — СЛЕПОК: при архивации сохраняются рассчитанные данные всех
виджетов (jsonb). Архив показывает «как было на момент архивации», даже если
исходные данные позже удалены ретенцией или сам дашборд удалён. Слепок
объективен (полные данные, без строкового RLS) — поэтому доступ в раздел
выдаётся избирательно и только администратором/модератором.
"""
from __future__ import annotations

import io
import json
from datetime import date, timedelta
from typing import List, Optional

from openpyxl import Workbook

from ..audit import service as audit_svc
from ._base import ANNOTATION_TYPES, DashboardError
from ._widgetdata import compute_widget_data
from ._widgetexport import _row

PRIVILEGED = {"admin", "superadmin", "moderator", "senior_moderator"}


# ── Доступ ────────────────────────────────────────────────────────────────────

async def can_view_archive(conn, org_id, user: dict) -> bool:
    """Раздел видят привилегированные роли всегда; остальные — по допуску.
    Роли приходят из get_current_user; фолбэк на запрос к БД оставлен для
    вызовов с «сырым» user (фоновые задачи, тесты)."""
    roles = user.get("roles")
    if roles is None:
        rows = await conn.fetch(
            "select r.code from user_roles ur join roles r on r.id=ur.role_id where ur.user_id=$1",
            user["id"])
        roles = [r["code"] for r in rows]
    if PRIVILEGED & set(roles):
        return True
    return bool(await conn.fetchval(
        "select 1 from archive_access where organization_id=$1 and user_id=$2",
        org_id, user["id"]))


async def _ensure_view(conn, org_id, user: dict) -> None:
    if not await can_view_archive(conn, org_id, user):
        raise DashboardError("Нет доступа к архиву")


async def list_access(conn, org_id) -> List[dict]:
    rows = await conn.fetch(
        "select a.user_id, u.login, u.full_name, a.granted_at "
        "from archive_access a join users u on u.id=a.user_id "
        "where a.organization_id=$1 order by u.full_name, u.login", org_id)
    return [dict(r) | {"user_id": str(r["user_id"])} for r in rows]


async def add_access(conn, org_id, actor_id, user_id: str) -> dict:
    u = await conn.fetchrow("select id from users where id=$1::uuid and organization_id=$2", user_id, org_id)
    if u is None:
        raise DashboardError("Пользователь не найден")
    await conn.execute(
        "insert into archive_access(organization_id, user_id, granted_by) values($1,$2::uuid,$3) "
        "on conflict do nothing", org_id, user_id, actor_id)
    await audit_svc.write_event(conn, org_id, actor_id, "grant_access", "archive", user_id,
                                new_data={"section": "archive", "user_id": user_id})
    return {"user_id": user_id}


async def remove_access(conn, org_id, actor_id, user_id: str) -> None:
    await conn.execute(
        "delete from archive_access where organization_id=$1 and user_id=$2::uuid", org_id, user_id)
    await audit_svc.write_event(conn, org_id, actor_id, "revoke_access", "archive", user_id,
                                old_data={"section": "archive", "user_id": user_id})


# ── Слепок ───────────────────────────────────────────────────────────────────

async def _build_snapshot(conn, org_id, dashboard_id: str) -> dict:
    """Рассчитать и упаковать данные всех страниц/виджетов дашборда.
    Объективный слепок: без строкового RLS (user=None) и без фильтров периода."""
    pages = await conn.fetch(
        "select id, name from dashboard_pages where dashboard_id=$1::uuid order by position, created_at",
        dashboard_id)
    out_pages = []
    for p in pages:
        widgets = await conn.fetch(
            "select id, name, widget_type, position_x, position_y, width, height "
            "from widgets where page_id=$1::uuid order by position_y, position_x", p["id"])
        out_w = []
        for w in widgets:
            item = {"id": str(w["id"]), "name": w["name"], "widget_type": w["widget_type"],
                    "x": w["position_x"], "y": w["position_y"], "w": w["width"], "h": w["height"]}
            try:
                item["data"] = await compute_widget_data(conn, org_id, str(w["id"]), skip_acl=True)
            except DashboardError as e:
                item["error"] = str(e)
            out_w.append(item)
        out_pages.append({"name": p["name"], "widgets": out_w})
    return {"pages": out_pages}


async def _dashboard_org(conn, org_id, dashboard_id: str):
    return await conn.fetchrow(
        "select id, name, publication_status, auto_archive from dashboards "
        "where id=$1::uuid and organization_id=$2", dashboard_id, org_id)


async def archive_dashboard(conn, org_id, user: dict, dashboard_id: str,
                            topic: Optional[str], note: Optional[str]) -> dict:
    """Ручная архивация: слепок + перевод дашборда в статус «в архиве»."""
    d = await _dashboard_org(conn, org_id, dashboard_id)
    if d is None:
        raise DashboardError("Дашборд не найден")
    if d["publication_status"] == "archived":
        raise DashboardError("Дашборд уже в архиве")
    snapshot = await _build_snapshot(conn, org_id, dashboard_id)
    month = date.today().strftime("%Y-%m")
    row = await conn.fetchrow(
        "insert into dashboard_archive(organization_id, dashboard_id, dashboard_name, topic, note, "
        "archive_month, snapshot, prev_status, auto, archived_by) "
        "values($1,$2::uuid,$3,$4,$5,$6,$7::jsonb,$8,false,$9) returning id, archived_at",
        org_id, dashboard_id, d["name"], (topic or "").strip() or None, (note or "").strip() or None,
        month, json.dumps(snapshot, ensure_ascii=False, default=str), d["publication_status"], user["id"])
    await conn.execute(
        "update dashboards set publication_status='archived', updated_at=now() where id=$1::uuid", dashboard_id)
    await audit_svc.write_event(conn, org_id, user["id"], "archive", "dashboard", dashboard_id,
                                new_data={"archive_id": str(row["id"]), "topic": topic, "month": month})
    return {"id": str(row["id"]), "archive_month": month}


async def unarchive(conn, org_id, user: dict, archive_id: str) -> dict:
    """Вернуть дашборд из архива: восстановить прежний статус. Слепок остаётся
    как историческая запись (его можно удалить отдельно, только администратор)."""
    a = await conn.fetchrow(
        "select id, dashboard_id, prev_status from dashboard_archive "
        "where id=$1::uuid and organization_id=$2", archive_id, org_id)
    if a is None:
        raise DashboardError("Запись архива не найдена")
    if a["dashboard_id"] is None:
        raise DashboardError("Исходный дашборд удалён — возвращать нечего (слепок остаётся)")
    prev = a["prev_status"] if a["prev_status"] in ("draft", "review", "published") else "draft"
    await conn.execute(
        "update dashboards set publication_status=$2, updated_at=now() where id=$1::uuid",
        a["dashboard_id"], prev)
    await audit_svc.write_event(conn, org_id, user["id"], "unarchive", "dashboard", str(a["dashboard_id"]),
                                new_data={"archive_id": archive_id, "restored_status": prev})
    return {"dashboard_id": str(a["dashboard_id"]), "publication_status": prev}


async def delete_archive(conn, org_id, user: dict, archive_id: str) -> None:
    """Удаление слепка — только администратор (гейт в роутере)."""
    a = await conn.fetchrow(
        "select id, dashboard_name, archive_month from dashboard_archive "
        "where id=$1::uuid and organization_id=$2", archive_id, org_id)
    if a is None:
        raise DashboardError("Запись архива не найдена")
    await conn.execute("delete from dashboard_archive where id=$1::uuid", archive_id)
    await audit_svc.write_event(conn, org_id, user["id"], "delete", "archive", archive_id,
                                old_data={"dashboard_name": a["dashboard_name"], "month": a["archive_month"]})


async def set_auto_archive(conn, org_id, user: dict, dashboard_id: str, enabled: bool) -> dict:
    d = await _dashboard_org(conn, org_id, dashboard_id)
    if d is None:
        raise DashboardError("Дашборд не найден")
    await conn.execute("update dashboards set auto_archive=$2 where id=$1::uuid", dashboard_id, enabled)
    return {"dashboard_id": dashboard_id, "auto_archive": enabled}


# ── Чтение архива ────────────────────────────────────────────────────────────

async def months(conn, org_id, user: dict) -> List[dict]:
    await _ensure_view(conn, org_id, user)
    rows = await conn.fetch(
        "select archive_month, count(*) as cnt from dashboard_archive "
        "where organization_id=$1 group by archive_month order by archive_month desc", org_id)
    return [{"month": r["archive_month"], "count": r["cnt"]} for r in rows]


async def topics(conn, org_id, user: dict) -> List[str]:
    await _ensure_view(conn, org_id, user)
    rows = await conn.fetch(
        "select distinct topic from dashboard_archive "
        "where organization_id=$1 and topic is not null order by topic", org_id)
    return [r["topic"] for r in rows]


async def list_archive(conn, org_id, user: dict, month: Optional[str] = None,
                       q: Optional[str] = None, topic: Optional[str] = None,
                       from_date: Optional[str] = None, to_date: Optional[str] = None) -> List[dict]:
    await _ensure_view(conn, org_id, user)
    where = "a.organization_id=$1"
    params: list = [org_id]
    if month:
        params.append(month); where += f" and a.archive_month=${len(params)}"
    if q and q.strip():
        params.append(f"%{q.strip()}%")
        where += (f" and (a.dashboard_name ilike ${len(params)} or a.topic ilike ${len(params)} or exists ("
                  f"select 1 from jsonb_array_elements(a.snapshot->'pages') pg "
                  f"where pg->>'name' ilike ${len(params)}))")
    if topic:
        params.append(topic); where += f" and a.topic=${len(params)}"
    if from_date:
        params.append(from_date); where += f" and a.archived_at::date >= ${len(params)}::text::date"
    if to_date:
        params.append(to_date); where += f" and a.archived_at::date <= ${len(params)}::text::date"
    rows = await conn.fetch(
        "select a.id, a.dashboard_id, a.dashboard_name, a.topic, a.note, a.archive_month, "
        "a.auto, a.archived_at, u.full_name as archived_by_name, "
        "jsonb_array_length(a.snapshot->'pages') as pages "
        "from dashboard_archive a left join users u on u.id=a.archived_by "
        f"where {where} order by a.archived_at desc limit 500", *params)
    return [dict(r) | {"id": str(r["id"]),
                       "dashboard_id": str(r["dashboard_id"]) if r["dashboard_id"] else None}
            for r in rows]


async def get_archive(conn, org_id, user: dict, archive_id: str) -> dict:
    await _ensure_view(conn, org_id, user)
    a = await conn.fetchrow(
        "select a.id, a.dashboard_id, a.dashboard_name, a.topic, a.note, a.archive_month, a.auto, "
        "a.archived_at, a.snapshot, u.full_name as archived_by_name "
        "from dashboard_archive a left join users u on u.id=a.archived_by "
        "where a.id=$1::uuid and a.organization_id=$2", archive_id, org_id)
    if a is None:
        raise DashboardError("Запись архива не найдена")
    out = dict(a) | {"id": str(a["id"]),
                     "dashboard_id": str(a["dashboard_id"]) if a["dashboard_id"] else None}
    out["snapshot"] = json.loads(a["snapshot"])
    return out


# ── Экспорт слепка в Excel ───────────────────────────────────────────────────

def _sheet_name(wb: Workbook, name: str) -> str:
    base = "".join(ch for ch in name if ch not in "[]:*?/\\")[:28] or "Лист"
    n, cand = 1, "".join(ch for ch in name if ch not in "[]:*?/\\")[:28] or "Лист"
    while cand in wb.sheetnames:
        n += 1; cand = f"{base[:25]}_{n}"
    return cand


def snapshot_to_xlsx(archive: dict) -> bytes:
    """Excel из ЗАМОРОЖЕННЫХ данных слепка (никаких перерасчётов)."""
    wb = Workbook()
    summary = wb.active
    summary.title = "Сводка"
    summary.append(["Дашборд", archive["dashboard_name"]])
    summary.append(["Тема", archive.get("topic") or "—"])
    summary.append(["Месяц архива", archive["archive_month"]])
    summary.append(["Архивировано", str(archive["archived_at"])[:19]])
    summary.append([])
    summary.append(["Страница", "Виджет", "Тип", "Показатель", "Значение"])
    for page in archive["snapshot"].get("pages", []):
        for w in page.get("widgets", []):
            d, t, name = w.get("data") or {}, w.get("widget_type"), w.get("name")
            if t in ANNOTATION_TYPES or not d:
                continue
            if t in ("kpi", "gauge"):
                summary.append([page["name"], name, t, "значение", d.get("value")])
            elif t == "plan_fact":
                summary.append([page["name"], name, t, "план", d.get("plan")])
                summary.append([page["name"], name, t, "факт", d.get("fact")])
                summary.append([page["name"], name, t, "выполнение, %", d.get("pct")])
            elif t in ("bar", "line", "pie", "waterfall", "objects_compare"):
                ws = wb.create_sheet(_sheet_name(wb, name))
                ws.append(["Категория", "Значение"])
                for c, v in zip(d.get("categories", []), d.get("values", []), strict=False):
                    ws.append([c, v])
            elif t == "dynamics":
                ws = wb.create_sheet(_sheet_name(wb, name))
                ws.append(["Период", "Значение"])
                for pr, v in zip(d.get("periods", []), d.get("values", []), strict=False):
                    # Дата — датой, а не строкой «2026-07-22»: слепок читают тем
                    # же Excel, что и обычную выгрузку (см. _widgetexport._row).
                    _row(ws, [pr, v])
            elif t == "yoy":
                ws = wb.create_sheet(_sheet_name(wb, name))
                py, cy = d.get("previous_year"), d.get("current_year")
                ws.append(["Месяц", str(py) if py else "пред. год", str(cy)])
                for mn, pv, cv in zip(d.get("months", []), d.get("previous", []), d.get("current", []), strict=False):
                    ws.append([mn, pv, cv])
            elif t == "compare":
                ws = wb.create_sheet(_sheet_name(wb, name))
                series = d.get("series", [])
                ws.append(["Категория"] + [s.get("name") for s in series])
                for i, c in enumerate(d.get("categories", [])):
                    ws.append([c] + [(s.get("data") or [None] * (i + 1))[i] if i < len(s.get("data", [])) else None
                                     for s in series])
            elif t == "table":
                ws = wb.create_sheet(_sheet_name(wb, name))
                cols = list(d.get("columns", []))
                # Имена показателей, а не коды полей — как на экране. У старых
                # слепков поля column_titles нет, поэтому запасной вариант — код.
                titles = d.get("column_titles") or {}
                ws.append(["Строка"] + [titles.get(c, c) for c in cols])
                for r in d.get("rows", []):
                    ws.append([r.get("row")] + [r.get(c) for c in cols])
            elif t == "pivot":
                ws = wb.create_sheet(_sheet_name(wb, name))
                cols = list(d.get("columns", []))
                ws.append(["Строка"] + cols + ["Итого"])
                for r in d.get("rows", []):
                    ws.append([r.get("row")] + list(r.get("values", [])) + [r.get("total")])
            elif t == "heatmap":
                ws = wb.create_sheet(_sheet_name(wb, name))
                cols, rws = list(d.get("columns", [])), list(d.get("rows", []))
                grid = [[None] * len(cols) for _ in rws]
                for ci, ri, v in d.get("cells", []):
                    if ri < len(rws) and ci < len(cols):
                        grid[ri][ci] = v
                ws.append(["Строка"] + cols)
                for i, rname in enumerate(rws):
                    ws.append([rname] + grid[i])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Ежемесячная автоархивация (вызывается планировщиком arq) ────────────────

async def run_monthly_auto_archive(conn, org_id) -> int:
    """Слепки всех дашбордов с флажком auto_archive за ПРОШЕДШИЙ месяц.
    Идемпотентно: если авто-слепок дашборда за месяц уже есть — пропуск."""
    month = (date.today().replace(day=1) - timedelta(days=1)).strftime("%Y-%m")
    rows = await conn.fetch(
        "select d.id, d.name, d.publication_status from dashboards d "
        "where d.organization_id=$1 and d.auto_archive "
        "and not exists (select 1 from dashboard_archive a "
        "  where a.dashboard_id=d.id and a.archive_month=$2 and a.auto)", org_id, month)
    n = 0
    for d in rows:
        snapshot = await _build_snapshot(conn, org_id, str(d["id"]))
        await conn.execute(
            "insert into dashboard_archive(organization_id, dashboard_id, dashboard_name, topic, note, "
            "archive_month, snapshot, prev_status, auto, archived_by) "
            "values($1,$2::uuid,$3,$4,$5,$6,$7::jsonb,$8,true,null)",
            org_id, str(d["id"]), d["name"], "Автоархив",
            f"Ежемесячный слепок за {month}", month,
            json.dumps(snapshot, ensure_ascii=False, default=str), d["publication_status"])
        n += 1
    return n
