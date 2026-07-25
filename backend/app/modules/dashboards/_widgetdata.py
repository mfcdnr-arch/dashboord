"""Вычисление данных виджета (вынесено из service.py).

Метрики (лучшая версия), датасеты (активный выпуск / по периодам / таблица),
сборка данных по типу виджета (_compute_widget), кэш (compute_widget_data),
предпросмотр, сработавшие алерты организации, xlsx-экспорт страницы, drill
(прозрачность показателя). Слой данных: зависит от _alerts/_base/_rls и метрик.
"""
from __future__ import annotations

import json
from typing import Dict, List, Optional

from ... import cache
from ..metrics import resolver as mr
from ..metrics.parser import FormulaError, extract_dependencies, parse
from ._alerts import _cfg, evaluate_alert
from ._base import WIDGET_TYPES, DashboardError
from ._rls import _can_view, visible_dashboard_ids


def _apply_target(res: dict, cfg: dict, value) -> None:
    """Цель/бенчмарк на показателе (KPI/gauge): добавляет target и % достижения."""
    t = cfg.get("target")
    if t is None or value is None:
        return
    try:
        tgt = float(t)
    except (TypeError, ValueError):
        return
    res["target"] = tgt
    res["target_label"] = cfg.get("target_label") or "Цель"
    res["target_pct"] = (float(value) / tgt * 100.0) if tgt else None


def _linear_trend(values: list) -> Optional[dict]:
    """Линейная регрессия y=a+b·x (x=0..n-1) по ряду значений (без ИИ, метод
    наименьших квадратов). Возвращает наклон и концы прямой для наложения."""
    ys = [v for v in values if v is not None]
    n = len(ys)
    if n < 2:
        return None
    xs = range(n)
    sx, sy = sum(xs), sum(ys)
    sxx = sum(x * x for x in xs)
    sxy = sum(x * ys[x] for x in xs)
    denom = n * sxx - sx * sx
    if denom == 0:
        return None
    b = (n * sxy - sx * sy) / denom
    a = (sy - b * sx) / n
    return {"slope": b, "endpoints": [a, a + b * (len(values) - 1)]}


async def _widget_org(conn, org_id, widget_id: str):
    return await conn.fetchrow(
        "select w.* from widgets w where w.id=$1::uuid and w.organization_id=$2", widget_id, org_id)


async def _page_org(conn, org_id, page_id: str):
    return await conn.fetchrow(
        "select p.id, p.dashboard_id from dashboard_pages p join dashboards d on d.id=p.dashboard_id "
        "where p.id=$1::uuid and d.organization_id=$2", page_id, org_id)


async def _best_metric_version(conn, org_id, code: str):
    # приоритет версии: одобренная → проверенная → любая (черновик не берётся вперёд проверенной)
    return await conn.fetchrow(
        "select m.name, m.info_text, m.description, mv.formula_expression, mv.formula_ast, mv.unit, mv.version_no, mv.status "
        "from metrics m join metric_versions mv on mv.metric_id=m.id "
        "where m.organization_id=$1 and m.code=$2 "
        "order by (case mv.status when 'approved' then 0 when 'validated' then 1 else 2 end), "
        "mv.version_no desc limit 1",
        org_id, code,
    )


async def _formula_value(conn, org_id, formula: str):
    """Вычисление произвольной формулы виджета (без сохранённой метрики)."""
    try:
        ast = parse(formula)
        return await mr.evaluate_ast(conn, org_id, ast)
    except FormulaError as e:
        raise DashboardError(str(e))


async def _metric_value(conn, org_id, code: str):
    row = await _best_metric_version(conn, org_id, code)
    if row is None:
        raise DashboardError(f"Метрика '{code}' не найдена")
    ast = row["formula_ast"]
    if isinstance(ast, str):
        ast = json.loads(ast)
    try:
        value = await mr.evaluate_ast(conn, org_id, ast)
    except FormulaError as e:
        raise DashboardError(str(e))
    return value, row["unit"]


async def _dataset_series(conn, org_id, dataset_code: str, value_field: str, row=None):
    rel = await mr._active_release(conn, org_id, dataset_code)
    if rel is None:
        raise DashboardError(f"Датасет '{dataset_code}' не найден или не выпущен")
    rows = await conn.fetch(
        "select row_label, value_number from dataset_values "
        "where dataset_release_id=$1 and canonical_field_code=$2 and value_number is not null "
        "and ($3::text is null or row_label=$3) order by row_index", rel, value_field, row,
    )
    return [{"category": r["row_label"], "value": float(r["value_number"])} for r in rows]


async def _dataset_multi_series(conn, org_id, dataset_code: str, value_fields: List[str], row=None) -> dict:
    """Несколько серий по одному датасету: категории=строки, серия=каждое поле."""
    rel = await mr._active_release(conn, org_id, dataset_code)
    if rel is None:
        raise DashboardError(f"Датасет '{dataset_code}' не найден или не выпущен")
    names = {r["code"]: r["name"] for r in await conn.fetch(
        "select drf.canonical_field_code as code, coalesce(cf.name, drf.canonical_field_code) as name "
        "from dataset_release_fields drf "
        "left join canonical_fields cf on cf.code=drf.canonical_field_code "
        "  and cf.object_id=(select object_id from dataset_releases where id=$1) "
        "where drf.dataset_release_id=$1", rel)}
    rows = await conn.fetch(
        "select row_index, row_label, canonical_field_code, value_number from dataset_values "
        "where dataset_release_id=$1 and canonical_field_code = any($2::text[]) and value_number is not null "
        "and ($3::text is null or row_label=$3) order by row_index", rel, value_fields, row)
    categories: List[str] = []
    per: Dict[str, Dict[str, float]] = {f: {} for f in value_fields}
    for r in rows:
        lbl = r["row_label"]
        if lbl not in categories:
            categories.append(lbl)
        per[r["canonical_field_code"]][lbl] = float(r["value_number"])
    series = [{"name": names.get(f, f), "data": [per[f].get(l) for l in categories]} for f in value_fields]
    return {"categories": categories, "series": series}


async def _dataset_period_series(conn, org_id, dataset_code: str, value_field: str,
                                from_date=None, to_date=None, row=None):
    """Ряд по периодам: для каждого активного выпуска датасета — сумма поля (динамика)."""
    rels = await conn.fetch(
        "select id, reporting_period_start from dataset_releases "
        "where organization_id=$1 and code=$2 and status <> 'superseded' "
        "and ($3::text is null or reporting_period_start >= $3::text::date) "
        "and ($4::text is null or reporting_period_start <= $4::text::date) "
        "order by reporting_period_start nulls last",
        org_id, dataset_code, from_date, to_date,
    )
    if not rels:
        raise DashboardError(f"Датасет '{dataset_code}' не найден или не выпущен")
    out = []
    for r in rels:
        s = await conn.fetchval(
            "select coalesce(sum(value_number),0) from dataset_values "
            "where dataset_release_id=$1 and canonical_field_code=$2 and ($3::text is null or row_label=$3)",
            r["id"], value_field, row)
        period = r["reporting_period_start"].isoformat() if r["reporting_period_start"] else "—"
        out.append((period, float(s)))
    return out


async def _dataset_table(conn, org_id, dataset_code: str, row=None):
    rel = await mr._active_release(conn, org_id, dataset_code)
    if rel is None:
        raise DashboardError(f"Датасет '{dataset_code}' не найден или не выпущен")
    fields = await conn.fetch(
        "select distinct canonical_field_code from dataset_values where dataset_release_id=$1 "
        "order by canonical_field_code", rel)
    cols = [f["canonical_field_code"] for f in fields]
    vals = await conn.fetch(
        "select row_index, row_label, canonical_field_code, value_text, value_number "
        "from dataset_values where dataset_release_id=$1 and ($2::text is null or row_label=$2) "
        "order by row_index", rel, row)
    by_row: Dict[int, dict] = {}
    for v in vals:
        r = by_row.setdefault(v["row_index"], {"__row__": v["row_label"]})
        r[v["canonical_field_code"]] = (
            float(v["value_number"]) if v["value_number"] is not None else v["value_text"])
    rows = [{"row": by_row[i].get("__row__"), **{c: by_row[i].get(c) for c in cols}}
            for i in sorted(by_row)]
    return {"columns": cols, "rows": rows}


async def compute_widget_data(conn, org_id, widget_id: str, from_date=None, to_date=None, row=None,
                              user: dict = None, skip_acl: bool = False) -> dict:
    w = await _widget_org(conn, org_id, widget_id)
    if w is None:
        raise DashboardError("Виджет не найден")
    # RLS: обход разрешён только явным skip_acl (внутренние агрегаты, где
    # видимость уже проверена выше). По HTTP всегда приходит user — fail closed.
    if not skip_acl and (user is None or not await _can_view(conn, org_id, user, str(w["dashboard_id"]))):
        raise DashboardError("Виджет не найден")
    # TTL-кэш: данные виджета одинаковы для всех, кто его видит (доступ проверен
    # выше). Ключ учитывает фильтры страницы. Мягкая деградация при недоступном Redis.
    key = f"wd:{widget_id}:{from_date or ''}:{to_date or ''}:{row or ''}"
    cached = await cache.get(key)
    if cached is not None:
        try:
            return json.loads(cached)
        except ValueError:
            pass
    result = await _compute_widget(conn, org_id, w["widget_type"], w["name"], _cfg(w), from_date, to_date, row)
    try:
        await cache.set(key, json.dumps(result, ensure_ascii=False), cache.WIDGET_DATA_TTL)
    except (TypeError, ValueError):
        pass  # несериализуемый результат не кэшируем
    return result


async def preview_widget(conn, org_id, widget_type: str, name: Optional[str], config: dict) -> dict:
    """Предпросмотр виджета по конфигу без сохранения (для конструктора)."""
    if widget_type not in WIDGET_TYPES:
        raise DashboardError(f"Неизвестный тип виджета: {widget_type}")
    return await _compute_widget(conn, org_id, widget_type, name or "Предпросмотр", config or {})


async def _compute_widget(conn, org_id, t: str, name: str, cfg: dict,
                          from_date=None, to_date=None, row=None) -> dict:
    # Виджетный фильтр (переопределение глобального): если у виджета задан
    # собственный фильтр (filter_scope='own'), он игнорирует фильтр страницы.
    if cfg.get("filter_scope") == "own":
        from_date = cfg.get("own_from") or None
        to_date = cfg.get("own_to") or None
        row = cfg.get("own_row") or None

    if t == "text":
        return {"type": "text", "title": name, "heading": cfg.get("heading"),
                "body": cfg.get("body"), "align": cfg.get("align", "left")}

    if t == "image":
        return {"type": "image", "title": name, "url": cfg.get("url"),
                "caption": cfg.get("caption"), "fit": cfg.get("fit", "contain")}

    if t == "compare":
        fields = cfg.get("value_fields") or []
        if not cfg.get("dataset_code") or not fields:
            raise DashboardError("Сравнение: укажите dataset_code и value_fields")
        res = await _dataset_multi_series(conn, org_id, cfg["dataset_code"], fields, row)
        res["type"], res["viz"], res["title"] = "compare", cfg.get("viz", "bar"), name
        return res

    if t == "heatmap":
        # Тепловая карта: матрица строки(датасета) × поля, значение — интенсивность цвета.
        # Для МФЦ удобно: услуги × периоды/отделы, нагрузка по строкам и столбцам.
        fields = cfg.get("value_fields") or []
        if not cfg.get("dataset_code") or not fields:
            raise DashboardError("Тепловая карта: укажите dataset_code и value_fields")
        ms = await _dataset_multi_series(conn, org_id, cfg["dataset_code"], fields, row)
        # ms: {categories:[строки], series:[{name:поле, data:[значения по строкам]}]}
        cols = [s["name"] for s in ms["series"]]
        cells = []  # [col_idx, row_idx, value]
        nums = []
        for ci, s in enumerate(ms["series"]):
            for ri, v in enumerate(s["data"]):
                if v is not None:
                    cells.append([ci, ri, v])
                    nums.append(v)
        return {"type": "heatmap", "title": name, "rows": ms["categories"], "columns": cols,
                "cells": cells, "min": (min(nums) if nums else 0), "max": (max(nums) if nums else 0)}

    if t == "pivot":
        # Сводная таблица: строки × поля + итоги по строкам, столбцам и общий.
        # Для МФЦ: услуги × показатели с автоматическими суммами (отчётность).
        fields = cfg.get("value_fields") or []
        if not cfg.get("dataset_code") or not fields:
            raise DashboardError("Сводная таблица: укажите dataset_code и value_fields")
        ms = await _dataset_multi_series(conn, org_id, cfg["dataset_code"], fields, row)
        cols = [s["name"] for s in ms["series"]]
        col_totals = [0.0] * len(cols)
        grand = 0.0
        rows_out = []
        for ri, rlabel in enumerate(ms["categories"]):
            vals, rtotal = [], 0.0
            for ci, s in enumerate(ms["series"]):
                v = s["data"][ri]
                vals.append(v)
                if v is not None:
                    rtotal += v
                    col_totals[ci] += v
                    grand += v
            rows_out.append({"row": rlabel, "values": vals, "total": rtotal})
        return {"type": "pivot", "title": name, "columns": cols, "rows": rows_out,
                "col_totals": col_totals, "grand_total": grand}

    if t == "waterfall":
        # Водопад: вклад каждой строки в накопленный итог (нарастающим), финальный столбец «Итого».
        # Для МФЦ: из чего складывается общий объём (услуги → суммарно).
        if not cfg.get("dataset_code") or not cfg.get("value_field"):
            raise DashboardError("Водопад: укажите dataset_code и value_field")
        series = await _dataset_series(conn, org_id, cfg["dataset_code"], cfg["value_field"], row)
        cats = [s["category"] for s in series]
        vals = [s["value"] for s in series]
        return {"type": "waterfall", "title": name, "categories": cats, "values": vals,
                "total_label": cfg.get("total_label") or "Итого"}

    if t == "objects_compare":
        # Сравнение подразделений: показатель (поле) агрегируется по ОБЪЕКТАМ
        # (каждый объект = подразделение/филиал), берётся последний выпуск на объект.
        field = cfg.get("value_field")
        if not field:
            raise DashboardError("Сравнение подразделений: укажите показатель (поле)")
        rows = await conn.fetch(
            "with latest as ("
            "  select distinct on (object_id) id, object_id from dataset_releases "
            "  where organization_id=$1 and status<>'superseded' and object_id is not null "
            "  order by object_id, reporting_period_start desc nulls last, created_at desc) "
            "select o.name as obj, coalesce(sum(dv.value_number),0) as val "
            "from latest l join objects o on o.id=l.object_id "
            "join dataset_values dv on dv.dataset_release_id=l.id and dv.canonical_field_code=$2 "
            "group by o.name having coalesce(sum(dv.value_number),0) <> 0 order by val desc",
            org_id, field)
        return {"type": "objects_compare", "title": name,
                "categories": [r["obj"] for r in rows], "values": [float(r["val"]) for r in rows]}

    if t == "dynamics":
        if not cfg.get("dataset_code") or not cfg.get("value_field"):
            raise DashboardError("Динамика: укажите dataset_code и value_field")
        series = await _dataset_period_series(conn, org_id, cfg["dataset_code"], cfg["value_field"], from_date, to_date, row)
        periods = [p for p, _ in series]
        values = [v for _, v in series]
        change = values[-1] - values[-2] if len(values) >= 2 else None
        change_pct = (change / values[-2] * 100.0) if (change is not None and values[-2]) else None
        res = {"type": "dynamics", "title": name, "periods": periods, "values": values,
               "change": change, "change_pct": change_pct}
        if cfg.get("trend"):
            tr = _linear_trend(values)
            if tr:
                res["trend"], res["trend_slope"] = tr["endpoints"], tr["slope"]
        res["alert"] = evaluate_alert("dynamics", cfg, res)
        return res

    if t == "kpi":
        if cfg.get("formula"):
            value, unit = await _formula_value(conn, org_id, cfg["formula"]), cfg.get("unit")
        elif cfg.get("metric_code"):
            value, unit = await _metric_value(conn, org_id, cfg["metric_code"])
        elif cfg.get("dataset_code") and cfg.get("value_field"):
            series = await _dataset_series(conn, org_id, cfg["dataset_code"], cfg["value_field"], row)
            value, unit = sum(s["value"] for s in series), cfg.get("unit")
        else:
            raise DashboardError("KPI: укажите формулу, metric_code или dataset_code+value_field")
        res = {"type": "kpi", "value": value, "unit": unit, "title": name}
        _apply_target(res, cfg, value)
        res["alert"] = evaluate_alert("kpi", cfg, res)
        return res

    if t == "gauge":
        # Спидометр: значение как у KPI + шкала (max). Идеален для «% выполнения».
        if cfg.get("formula"):
            value, unit = await _formula_value(conn, org_id, cfg["formula"]), cfg.get("unit")
        elif cfg.get("metric_code"):
            value, unit = await _metric_value(conn, org_id, cfg["metric_code"])
        elif cfg.get("dataset_code") and cfg.get("value_field"):
            series = await _dataset_series(conn, org_id, cfg["dataset_code"], cfg["value_field"], row)
            value, unit = sum(s["value"] for s in series), cfg.get("unit")
        else:
            raise DashboardError("Gauge: укажите формулу, metric_code или dataset_code+value_field")
        gmax = cfg.get("gauge_max")
        if gmax is None:
            gmax = 100 if (unit and "%" in unit) else (round((value or 0) * 1.25) or 100)
        res = {"type": "gauge", "value": value, "unit": unit, "max": gmax, "title": name}
        _apply_target(res, cfg, value)
        res["alert"] = evaluate_alert("kpi", cfg, res)  # те же пороги, что и KPI
        return res

    if t == "plan_fact":
        if cfg.get("plan_metric") and cfg.get("fact_metric"):
            plan, unit = await _metric_value(conn, org_id, cfg["plan_metric"])
            fact, _ = await _metric_value(conn, org_id, cfg["fact_metric"])
        elif cfg.get("dataset_code") and cfg.get("plan_field") and cfg.get("fact_field"):
            plan = sum(s["value"] for s in await _dataset_series(conn, org_id, cfg["dataset_code"], cfg["plan_field"], row))
            fact = sum(s["value"] for s in await _dataset_series(conn, org_id, cfg["dataset_code"], cfg["fact_field"], row))
            unit = cfg.get("unit")
        else:
            raise DashboardError("План-факт: укажите plan_metric+fact_metric или dataset_code+plan_field+fact_field")
        pct = (fact / plan * 100.0) if plan else None
        res = {"type": "plan_fact", "plan": plan, "fact": fact, "delta": fact - plan, "pct": pct, "unit": unit, "title": name}
        res["alert"] = evaluate_alert("plan_fact", cfg, res)
        return res

    if t == "table":
        if not cfg.get("dataset_code"):
            raise DashboardError("Таблица: укажите dataset_code")
        table = await _dataset_table(conn, org_id, cfg["dataset_code"], row)
        return {"type": "table", "title": name, **table}

    # bar | line | pie
    if not cfg.get("dataset_code") or not cfg.get("value_field"):
        raise DashboardError("График: укажите dataset_code и value_field")
    series = await _dataset_series(conn, org_id, cfg["dataset_code"], cfg["value_field"], row)
    return {"type": t, "title": name,
            "categories": [s["category"] for s in series],
            "values": [s["value"] for s in series]}


async def list_org_alerts(conn, org_id, user: dict, limit: int = 30) -> List[dict]:
    """Сработавшие KPI-алерты по доступным пользователю дашбордам — для «Главной».
    Возвращаются только уровни warn/danger (good — позитивная подсветка, не тревога)."""
    visible = await visible_dashboard_ids(conn, org_id, user)
    if not visible:
        return []
    rows = await conn.fetch(
        "select w.id, w.name, w.widget_type, w.dashboard_id, "
        "d.name as dashboard_name, d.publication_status, "
        "(select p.name from dashboard_pages p where p.id=w.page_id) as page_name "
        "from widgets w join dashboards d on d.id=w.dashboard_id "
        "where w.organization_id=$1 and w.dashboard_id = any($2::uuid[]) "
        "and (w.config ->> 'alerts') is not null "
        "order by d.name", org_id, list(visible),
    )
    out: List[dict] = []
    for w in rows:
        try:
            data = await compute_widget_data(conn, org_id, str(w["id"]), skip_acl=True)
        except DashboardError:
            continue
        al = data.get("alert")
        if not al or al["level"] not in ("warn", "danger"):
            continue
        out.append({
            "widget_id": str(w["id"]), "widget_name": w["name"], "widget_type": w["widget_type"],
            "dashboard_id": str(w["dashboard_id"]), "dashboard_name": w["dashboard_name"],
            "page_name": w["page_name"], "published": w["publication_status"] == "published",
            "level": al["level"], "label": al["label"], "measure": al["measure"], "unit": data.get("unit"),
        })
    out.sort(key=lambda x: (0 if x["level"] == "danger" else 1, x["dashboard_name"]))
    return out[:limit]


async def export_page_xlsx(conn, org_id, user: dict, page_id: str) -> bytes:
    """Экспорт данных всех виджетов страницы в .xlsx (openpyxl).
    KPI/план-факт — на лист «Сводка», датасетные виджеты — по листу на виджет.
    Аннотации (text/image) пропускаются. RLS: проверяется доступ к дашборду."""
    import io
    import re
    from openpyxl import Workbook

    p = await _page_org(conn, org_id, page_id)
    if p is None:
        raise DashboardError("Страница не найдена")
    if not await _can_view(conn, org_id, user, str(p["dashboard_id"])):
        raise DashboardError("Страница не найдена")

    rows = await conn.fetch(
        "select id, name, widget_type from widgets where page_id=$1::uuid order by position_y, position_x", page_id)

    wb = Workbook()
    summary = wb.active
    summary.title = "Сводка"
    summary.append(["Виджет", "Тип", "Показатель", "Значение"])
    has_summary = False

    used: set = set()
    def sheet_name(base: str) -> str:
        n = re.sub(r"[\[\]:*?/\\]", " ", base or "Лист")[:28].strip() or "Лист"
        cand, i = n, 2
        while cand.lower() in used:
            cand, i = f"{n[:25]} {i}", i + 1
        used.add(cand.lower())
        return cand

    for w in rows:
        wid, t, name = str(w["id"]), w["widget_type"], w["name"]
        if t in ("text", "image"):
            continue
        try:
            data = await compute_widget_data(conn, org_id, wid, skip_acl=True)
        except DashboardError:
            continue
        if t == "kpi":
            summary.append([name, "KPI", "значение", data.get("value")]); has_summary = True
        elif t == "gauge":
            summary.append([name, "Спидометр", "значение", data.get("value")]); has_summary = True
        elif t == "plan_fact":
            summary.append([name, "План-факт", "план", data.get("plan")])
            summary.append([name, "План-факт", "факт", data.get("fact")])
            summary.append([name, "План-факт", "выполнение, %", data.get("pct")]); has_summary = True
        elif t == "table":
            ws = wb.create_sheet(sheet_name(name))
            cols = list(data.get("columns", []))
            ws.append(["Строка"] + cols)
            for r in data.get("rows", []):
                ws.append([r.get("row")] + [r.get(c) for c in cols])
        elif t in ("bar", "line", "pie"):
            ws = wb.create_sheet(sheet_name(name))
            ws.append(["Категория", "Значение"])
            for c, v in zip(data.get("categories", []), data.get("values", [])):
                ws.append([c, v])
        elif t == "dynamics":
            ws = wb.create_sheet(sheet_name(name))
            ws.append(["Период", "Значение"])
            for pr, v in zip(data.get("periods", []), data.get("values", [])):
                ws.append([pr, v])
        elif t == "compare":
            ws = wb.create_sheet(sheet_name(name))
            series = data.get("series", [])
            cats = data.get("categories", [])
            ws.append(["Категория"] + [s.get("name") for s in series])
            for i, c in enumerate(cats):
                ws.append([c] + [(s.get("data") or [])[i] if i < len(s.get("data", [])) else None for s in series])
        elif t == "heatmap":
            ws = wb.create_sheet(sheet_name(name))
            cols = list(data.get("columns", []))
            rws = list(data.get("rows", []))
            grid = [[None] * len(cols) for _ in rws]
            for ci, ri, v in data.get("cells", []):
                if ri < len(rws) and ci < len(cols):
                    grid[ri][ci] = v
            ws.append(["Строка"] + cols)
            for i, rname in enumerate(rws):
                ws.append([rname] + grid[i])
        elif t == "pivot":
            ws = wb.create_sheet(sheet_name(name))
            cols = list(data.get("columns", []))
            ws.append(["Строка"] + cols + ["Итого"])
            for r in data.get("rows", []):
                ws.append([r.get("row")] + list(r.get("values", [])) + [r.get("total")])
            ws.append(["Итого"] + list(data.get("col_totals", [])) + [data.get("grand_total")])
        elif t == "waterfall":
            ws = wb.create_sheet(sheet_name(name))
            ws.append(["Категория", "Значение"])
            for c, v in zip(data.get("categories", []), data.get("values", [])):
                ws.append([c, v])
            ws.append([data.get("total_label", "Итого"), sum(v for v in data.get("values", []) if v is not None)])
        elif t == "objects_compare":
            ws = wb.create_sheet(sheet_name(name))
            ws.append(["Подразделение", "Значение"])
            for c, v in zip(data.get("categories", []), data.get("values", [])):
                ws.append([c, v])

    if not has_summary and len(wb.sheetnames) > 1:
        wb.remove(summary)  # нет KPI/план-факта, но есть датасетные листы — убираем пустую сводку

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def widget_drill(conn, org_id, widget_id: str, user: dict) -> dict:
    """Прозрачность показателя: из чего собран виджет — формулы метрик (уровень 1)
    и первичные строки датасетов (уровень 2)."""
    w = await _widget_org(conn, org_id, widget_id)
    if w is None:
        raise DashboardError("Виджет не найден")
    if not await _can_view(conn, org_id, user, str(w["dashboard_id"])):
        raise DashboardError("Виджет не найден")
    cfg = _cfg(w)
    metric_codes = [cfg[k] for k in ("metric_code", "plan_metric", "fact_metric") if cfg.get(k)]
    dataset_codes = [cfg["dataset_code"]] if cfg.get("dataset_code") else []

    metrics_info: List[dict] = []

    # Собственная формула виджета (KPI с config.formula) — показать как показатель уровня 1
    if cfg.get("formula"):
        try:
            deps = extract_dependencies(parse(cfg["formula"]))
        except FormulaError:
            deps = {"datasets": [], "metrics": []}
        metrics_info.append({
            "code": "(формула виджета)", "name": w["name"], "formula": cfg["formula"],
            "status": "widget", "version_no": None, "datasets": deps["datasets"],
        })
        dataset_codes += deps["datasets"]
        metric_codes += deps.get("metrics", [])

    for code in metric_codes:
        row = await _best_metric_version(conn, org_id, code)
        if row is None:
            continue
        ast = row["formula_ast"]
        if isinstance(ast, str):
            ast = json.loads(ast)
        deps = extract_dependencies(ast)
        metrics_info.append({
            "code": code, "name": row["name"], "formula": row["formula_expression"],
            "status": row["status"], "version_no": row["version_no"], "datasets": deps["datasets"],
            "info_text": row["info_text"], "description": row["description"],
        })
        dataset_codes += deps["datasets"]

    seen: List[str] = []
    for dc in dataset_codes:
        if dc not in seen:
            seen.append(dc)
    tables: Dict[str, Any] = {}
    for dc in seen:
        try:
            tables[dc] = await _dataset_table(conn, org_id, dc)
        except DashboardError:
            tables[dc] = {"columns": [], "rows": []}

    return {"widget": w["name"], "widget_type": w["widget_type"],
            "metrics": metrics_info, "datasets": seen, "tables": tables}
