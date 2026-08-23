"""Выгрузка ОДНОГО виджета (п. 7 списка предложений).

Раньше выгружалась только страница целиком, и человеку, которому нужна одна
таблица для доклада, приезжал файл на семнадцать листов. Главное, что
проверяем: один и тот же виджет ложится в файл ОДИНАКОВО, какой кнопкой его ни
выгружай (общий `_dump_widget`), и что фильтры страницы доезжают до файла —
иначе выгрузка разошлась бы с экраном.
"""
import io

import pytest

pytestmark = pytest.mark.asyncio(loop_scope="session")

from conftest import purge_dashboard


def _sheets(blob: bytes):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(blob))
    return {ws.title: [[c for c in row] for row in ws.iter_rows(values_only=True)] for ws in wb.worksheets}


async def _make(client, headers, seed_dataset, widget: dict):
    did = (await client.post("/dashboards", headers=headers, json={"name": "ztest_wexport"})).json()["id"]
    page = (await client.post(f"/dashboards/{did}/pages", headers=headers, json={"name": "Стр"})).json()
    w = (await client.post(f"/dashboard-pages/{page['id']}/widgets", headers=headers, json=widget)).json()
    return did, page["id"], w["id"]


async def test_one_widget_file_matches_the_page_file(client, admin_headers, seed_dataset):
    """Лист данных совпадает с тем, что даёт выгрузка всей страницы."""
    did, page_id, wid = await _make(client, admin_headers, seed_dataset, {
        "name": "Таблица района", "widget_type": "table",
        "config": {"dataset_code": seed_dataset["code"]}})
    try:
        one = _sheets((await client.get(f"/widgets/{wid}/export.xlsx", headers=admin_headers)).content)
        whole = _sheets((await client.get(f"/dashboard-pages/{page_id}/export.xlsx", headers=admin_headers)).content)
        # В файле одного виджета: «Содержание» + лист данных, и ничего лишнего.
        assert "Содержание" in one and len(one) == 2
        data_one = next(rows for title, rows in one.items() if title != "Содержание")
        data_all = next(rows for title, rows in whole.items() if title not in ("Содержание", "Сводка"))
        assert data_one == data_all, "один и тот же виджет должен лечь в файл одинаково"
        # «Содержание» отвечает, что это за файл: полное имя виджета.
        assert any("Таблица района" in str(c) for row in one["Содержание"] for c in row)
    finally:
        await purge_dashboard(did)


async def test_filters_reach_the_file(client, admin_headers, seed_dataset):
    """Файл обязан совпадать с экраном: фильтр строки сужает и его."""
    did, _page, wid = await _make(client, admin_headers, seed_dataset, {
        "name": "Таблица", "widget_type": "table",
        "config": {"dataset_code": seed_dataset["code"]}})
    try:
        full = _sheets((await client.get(f"/widgets/{wid}/export.xlsx", headers=admin_headers)).content)
        one_row = _sheets((await client.get(f"/widgets/{wid}/export.xlsx?row=ИНН",
                                            headers=admin_headers)).content)
        rows_full = next(r for t, r in full.items() if t != "Содержание")
        rows_filtered = next(r for t, r in one_row.items() if t != "Содержание")
        assert len(rows_full) - 1 == len(seed_dataset["rows"])
        assert len(rows_filtered) - 1 == 1 and rows_filtered[1][0] == "ИНН"
        # И сам фильтр назван в «Содержании» — иначе отфильтрованный файл не
        # отличить от полного.
        assert any("ИНН" in str(c) for row in one_row["Содержание"] for c in row)
    finally:
        await purge_dashboard(did)


async def test_annotation_has_nothing_to_export_and_stranger_gets_404(client, admin_headers, viewer, seed_dataset):
    did, _page, wid = await _make(client, admin_headers, seed_dataset, {
        "name": "Заголовок", "widget_type": "text", "config": {"heading": "Раздел"}})
    try:
        r = await client.get(f"/widgets/{wid}/export.xlsx", headers=admin_headers)
        assert r.status_code == 400 and "нет данных" in r.json()["detail"]
        # Чужой дашборд зрителю не виден — и выгрузка не должна становиться
        # обходным путём к его данным.
        assert (await client.get(f"/widgets/{wid}/export.xlsx", headers=viewer["headers"])).status_code == 404
    finally:
        await purge_dashboard(did)
