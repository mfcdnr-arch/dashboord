"""«Призрачная» линия прошлого отчёта на графиках (п. 3 второй волны).

График по строкам показывает СРЕЗ одного отчёта и сам по себе не отвечает на
вопрос «а было сколько?». Бледная серия позади текущей отвечает на месте — но
только если она про ТЕ ЖЕ строки. Отсюда состав проверок:

1. **Сопоставление по НАЗВАНИЮ строки, а не по номеру.** Между отчётами строки
   добавляют, убирают и переставляют; позиционное совпадение однажды подставило
   бы одному району цифру другого — молча и правдоподобно. Это главный тест.
2. **Закреплённый срез берёт отчёт, предшествующий ЕМУ**, а не последнему:
   иначе страница за июль показывала бы «было» из августа.
3. **RLS**: призрак читается тем же набором разрешённых строк, что и сам
   виджет, — иначе прошлый период показал бы строки, которых видеть нельзя.
4. **Молчать нельзя.** Нет прошлого отчёта или ни одна строка не сошлась —
   галочка включена, а на графике пусто; это выглядит поломкой, поэтому
   возвращается причина словами.
5. **Файл не расходится с экраном**: призрак уезжает в выгрузку колонкой.
"""
import pytest

pytestmark = pytest.mark.asyncio(loop_scope="session")

from conftest import db, purge_dashboard


async def _widget(client, headers, wtype, config, name="G"):
    did = (await client.post("/dashboards", headers=headers, json={"name": f"ztest_ghost_{name}"})).json()["id"]
    pid = (await client.post(f"/dashboards/{did}/pages", headers=headers, json={"name": "P"})).json()["id"]
    wid = (await client.post(f"/dashboard-pages/{pid}/widgets", headers=headers,
           json={"name": name, "widget_type": wtype, "config": config})).json()["id"]
    return did, pid, wid


async def test_ghost_matches_by_row_name_not_position(client, admin_headers, ids):
    """Прошлый отчёт раскладывается по НАЗВАНИЯМ строк.

    В прошлом отчёте строки идут в другом порядке и есть лишняя, которой
    сейчас нет. При позиционном сопоставлении «Горловка» получила бы чужое
    число — тест ловит именно это.
    """
    async with db.acquire() as conn:
        await _drop(conn, ids["org"])
        obj = await conn.fetchval(
            "insert into objects(organization_id,name) values($1,'ztest_gh_obj') returning id", ids["org"])
        old = await conn.fetchval(
            "insert into dataset_releases(organization_id,code,name,status,reporting_period_start,"
            "created_by,object_id) values($1,'ztest_gh_ds','ДС','released','2026-07-01',$2,$3) returning id",
            ids["org"], ids["admin"], obj)
        new = await conn.fetchval(
            "insert into dataset_releases(organization_id,code,name,status,reporting_period_start,"
            "created_by,object_id) values($1,'ztest_gh_ds','ДС','released','2026-07-08',$2,$3) returning id",
            ids["org"], ids["admin"], obj)
        # Прошлый отчёт: ДРУГОЙ порядок + строка «Ясиноватая», которой сейчас нет.
        for i, (lbl, v) in enumerate([("Горловка", 30), ("Ясиноватая", 999), ("Донецк", 10), ("Макеевка", 20)]):
            await conn.execute(
                "insert into dataset_values(dataset_release_id,row_index,row_label,canonical_field_code,"
                "value_number) values($1,$2,$3,'v',$4)", old, i, lbl, v)
        for i, (lbl, v) in enumerate([("Донецк", 11), ("Макеевка", 22), ("Горловка", 33)]):
            await conn.execute(
                "insert into dataset_values(dataset_release_id,row_index,row_label,canonical_field_code,"
                "value_number) values($1,$2,$3,'v',$4)", new, i, lbl, v)

    did, _, wid = await _widget(client, admin_headers, "bar",
                                {"dataset_code": "ztest_gh_ds", "value_field": "v", "ghost_prev": True})
    try:
        d = (await client.get(f"/widgets/{wid}/data", headers=admin_headers)).json()
        assert d["categories"] == ["Донецк", "Макеевка", "Горловка"]
        assert d["values"] == [11, 22, 33]
        assert d["ghost"]["period"] == "2026-07-01"
        # Ключевое: каждому городу — ЕГО прошлое значение, а не соседнее.
        assert d["ghost"]["values"] == [10, 20, 30], "сопоставление по имени, а не по порядку"
        assert 999 not in d["ghost"]["values"], "строки, которой сейчас нет, на графике быть не может"
        assert "ghost_note" not in d

        # Без галочки призрака нет вовсе — лишних запросов не делаем.
        await client.patch(f"/widgets/{wid}", headers=admin_headers,
                           json={"config": {"dataset_code": "ztest_gh_ds", "value_field": "v"}})
        d2 = (await client.get(f"/widgets/{wid}/data", headers=admin_headers)).json()
        assert "ghost" not in d2 and "ghost_note" not in d2
    finally:
        await purge_dashboard(did)
        async with db.acquire() as conn:
            await _drop(conn, ids["org"])


async def test_pinned_slice_looks_back_from_itself(client, admin_headers, seed_dataset):
    """Закреплённый срез сравнивается с отчётом, предшествующим ЕМУ.

    У `t_ds` два выпуска: 01.01 и 01.02. Срез за 01.02 обязан показать «было»
    за 01.01, а срез за 01.01 — сказать, что сравнивать не с чем, а не тянуть
    февральские цифры назад во времени.
    """
    did, _, wid = await _widget(client, admin_headers, "bar",
                                {"dataset_code": "t_ds", "value_field": "plan",
                                 "ghost_prev": True, "period": "2026-02-01"}, name="pin")
    try:
        d = (await client.get(f"/widgets/{wid}/data", headers=admin_headers)).json()
        assert d["values"] == [100, 50, 30]
        assert d["ghost"]["period"] == "2026-01-01"
        assert d["ghost"]["values"] == [95, 45, 25]

        # Самый первый отчёт: сравнивать не с чем — говорим словами.
        await client.patch(f"/widgets/{wid}", headers=admin_headers,
                           json={"config": {"dataset_code": "t_ds", "value_field": "plan",
                                            "ghost_prev": True, "period": "2026-01-01"}})
        first = (await client.get(f"/widgets/{wid}/data", headers=admin_headers)).json()
        assert "ghost" not in first
        assert "первый отчёт" in first["ghost_note"]
    finally:
        await purge_dashboard(did)


async def test_ghost_respects_row_rls(client, admin_headers, viewer, seed_dataset):
    """Призрак читается тем же набором разрешённых строк, что и сам виджет.

    Иначе «как было раньше» стало бы обходным путём к строкам, которые
    человеку видеть нельзя.
    """
    async with db.acquire() as conn:
        obj = str(await conn.fetchval("select id from objects where name='t_obj'"))
        org = await conn.fetchval("select organization_id from objects where id=$1::uuid", obj)
        dep = str(await conn.fetchval(
            "insert into departments(organization_id,name) values($1,'ztest_dep_ghost') returning id", org))
        await conn.execute("update users set department_id=$1::uuid where id=$2::uuid", dep, viewer["id"])

    did, _, wid = await _widget(client, admin_headers, "bar",
                                {"dataset_code": "t_ds", "value_field": "plan", "ghost_prev": True}, name="rls")
    try:
        await client.put(f"/objects/{obj}/row-acl/{dep}", headers=admin_headers,
                         json={"row_labels": ["Паспорт"]})
        await client.post(f"/dashboards/{did}/grants", headers=admin_headers,
                          json={"grantee_type": "user", "user_id": viewer["id"]})
        await client.post(f"/dashboards/{did}/publish", headers=admin_headers)

        d = (await client.get(f"/widgets/{wid}/data", headers=viewer["headers"])).json()
        assert d["categories"] == ["Паспорт"], "виджет показывает только разрешённую строку"
        assert d["ghost"]["values"] == [95], "и призрак тоже — только её"
        assert 45 not in d["ghost"]["values"] and 25 not in d["ghost"]["values"]

        # У администратора та же картина, но по всем строкам.
        a = (await client.get(f"/widgets/{wid}/data", headers=admin_headers)).json()
        assert a["ghost"]["values"] == [95, 45, 25]
    finally:
        await purge_dashboard(did)
        async with db.acquire() as conn:
            await conn.execute("delete from data_row_acl where object_id=$1::uuid", obj)
            await conn.execute("update users set department_id=null where id=$1::uuid", viewer["id"])
            await conn.execute("delete from departments where name='ztest_dep_ghost'")


async def test_compare_has_no_ghost_but_export_still_carries_it(client, admin_headers, seed_dataset):
    """У «Сравнения» призрака нет — сознательно; у столбцов он доезжает до файла.

    Призрак у «Сравнения» удваивал число столбиков, а `barGap` в ECharts
    действует на всю группу серий и совместить их попарно не может — выходил
    частокол вместо ответа (замерено на виджете заказчика: 13 показателей → 26
    полосок по 11px двумя отдельными кучами). Тест держит это решение, чтобы
    призрак не вернули туда «за компанию».

    Заодно проверяем правило проекта: то, что человек видит на экране, обязано
    быть и в выгрузке.
    """
    did, _, wid = await _widget(client, admin_headers, "compare",
                                {"dataset_code": "t_ds", "value_fields": ["plan"],
                                 "ghost_prev": True}, name="cmp")
    try:
        d = (await client.get(f"/widgets/{wid}/data", headers=admin_headers)).json()
        assert "ghost" not in d and "ghost_note" not in d, "у «Сравнения» призрака быть не должно"
    finally:
        await purge_dashboard(did)

    did2, _, wid2 = await _widget(client, admin_headers, "bar",
                                  {"dataset_code": "t_ds", "value_field": "plan",
                                   "ghost_prev": True}, name="exp")
    try:
        r = await client.get(f"/widgets/{wid2}/export.xlsx", headers=admin_headers)
        assert r.status_code == 200
        import io

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb[wb.sheetnames[-1]]
        head = [c.value for c in ws[1]]
        assert any(h and "Было 01.01.2026" in h for h in head), f"призрак обязан быть в файле: {head}"
        body = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
        assert body[0][1] == 100 and body[0][2] == 95, "и текущее, и прошлое значение рядом"
    finally:
        await purge_dashboard(did2)


async def _drop(conn, org):
    await conn.execute("delete from dataset_values where dataset_release_id in "
                       "(select id from dataset_releases where code like 'ztest_gh%')")
    await conn.execute("delete from dataset_releases where code like 'ztest_gh%'")
    await conn.execute("delete from objects where name like 'ztest_gh%' and organization_id=$1", org)
