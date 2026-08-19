"""Выгрузка журналов в CSV/XLSX (admin): аудит действий и журнал входов.

Волна B: доступ к аудиту/журналу входов для роли admin (не superadmin) теперь
требует явного гранта (audit_access_grants) — выдаём его сид-admin'у на время
этого файла (autouse), иначе все тесты ниже стали бы падать 403. Сам механизм
гранта/отзыва и негативные сценарии — в test_audit_access.py."""
import pytest
import pytest_asyncio

pytestmark = pytest.mark.asyncio(loop_scope="session")

from conftest import hdr, login

XLSX_SIG = b"PK\x03\x04"  # zip-сигнатура (xlsx — это zip)


@pytest_asyncio.fixture(autouse=True)
async def _grant_admin_audit_access(client, ids):
    sa = hdr(await login(client, "superadmin", "superadmin"))
    await client.post(f"/audit/access/{ids['admin']}", headers=sa)
    yield
    await client.delete(f"/audit/access/{ids['admin']}", headers=sa)


async def test_audit_csv(client, admin_headers):
    r = await client.get("/audit/export.csv", headers=admin_headers)
    assert r.status_code == 200
    assert "text/csv" in r.headers.get("content-type", "")
    assert r.content[:3] == b"\xef\xbb\xbf"  # BOM UTF-8


async def test_audit_xlsx(client, admin_headers):
    r = await client.get("/audit/export.xlsx", headers=admin_headers)
    assert r.status_code == 200
    assert r.content[:4] == XLSX_SIG


async def test_login_events_csv(client, admin_headers):
    r = await client.get("/login-events/export.csv", headers=admin_headers)
    assert r.status_code == 200
    assert r.content[:3] == b"\xef\xbb\xbf"


async def test_login_events_xlsx(client, admin_headers):
    r = await client.get("/login-events/export.xlsx", headers=admin_headers)
    assert r.status_code == 200
    assert r.content[:4] == XLSX_SIG


async def test_export_requires_admin(client, viewer):
    # непривилегированный (роль user) не должен выгружать журналы
    r = await client.get("/audit/export.csv", headers=viewer["headers"])
    assert r.status_code == 403


async def test_page_xlsx_sheets_dates_and_titles(client, admin_headers, ids, seed_dataset):
    """Файл страницы: листы различимы, даты — датами, заголовки — именами.

    Три свойства, каждое из которых уже ломалось: имена листов сливались в
    «… поль 2 / 3 / 4», период уезжал текстом «2026-01-01», а в таблице
    заголовками стояли машинные коды полей.
    """
    import io as _io

    from openpyxl import load_workbook

    from app import db
    from conftest import purge_dashboard

    did = (await client.post("/dashboards", headers=admin_headers,
                             json={"name": "ztest_xlsx_sheets"})).json()["id"]
    try:
        pid = (await client.post(f"/dashboards/{did}/pages", headers=admin_headers,
                                 json={"name": "P"})).json()["id"]
        # Имена нарочно длинные и почти одинаковые — как у граф госформы.
        long_name = ("Динамика: Количество обращений за результатом оказания услуг "
                     "в МФЦ · Факт · {}")
        async with db.acquire() as conn:
            await conn.execute(
                "insert into canonical_fields(object_id, code, name) "
                "select o.id, 'plan', 'План на период' from objects o "
                "where o.name='t_obj' and o.organization_id=$1 on conflict do nothing", ids["org"])
        for slice_ in ("нарастающим итогом**", "нарастающим итогом (текущий месяц)"):
            await client.post(f"/dashboard-pages/{pid}/widgets", headers=admin_headers, json={
                "name": long_name.format(slice_), "widget_type": "dynamics",
                "config": {"dataset_code": "t_ds", "value_field": "plan"}})
        await client.post(f"/dashboard-pages/{pid}/widgets", headers=admin_headers, json={
            "name": "Тест: таблица", "widget_type": "table",
            "config": {"dataset_code": "t_ds"}})

        r = await client.get(f"/dashboard-pages/{pid}/export.xlsx", headers=admin_headers)
        assert r.status_code == 200
        wb = load_workbook(_io.BytesIO(r.content))

        titles = wb.sheetnames
        assert titles[0] == "Содержание"                      # оглавление первым листом
        assert all(len(t) <= 31 for t in titles)              # жёсткий предел Excel
        assert len({t.lower() for t in titles}) == len(titles)  # листы различимы
        # Полные имена доступны целиком, как бы ни было урезано имя листа.
        toc = [(c[0].value, c[1].value) for c in wb["Содержание"].iter_rows(min_row=2, max_col=2)]
        assert any(full == long_name.format("нарастающим итогом**") for _, full in toc)
        assert {sheet for sheet, _ in toc} <= set(titles)     # оглавление не врёт

        dyn = wb[[s for s, f in toc if f.startswith("Динамика")][0]]
        first = dyn.cell(row=2, column=1)
        assert hasattr(first.value, "year"), "период должен быть датой, а не строкой"
        assert first.number_format == "DD.MM.YYYY"

        tbl = wb[[s for s, f in toc if f.startswith("Тест")][0]]
        assert "План на период" in [c.value for c in tbl[1]], "заголовок — имя, а не код поля"
    finally:
        await purge_dashboard(did)
