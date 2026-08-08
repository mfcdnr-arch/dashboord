"""Геометрия распознавания: объединения, баннер письма, уникальные коды полей.

Проверяется на форме, которая пришла от заказчика (МФЦ ДНР, 07.08.2026):
«приложение к письму» — сверху баннер, объединённый во всю ширину листа, ниже
многоэтажная шапка с объединениями, и только потом данные. На такой форме
прежний разбор давал всем 16 столбцам одно имя «ИНФОРМАЦИЯ», шапкой считал
текст письма, а выпуск датасета упал бы на unique-ограничении.

Тесты чистые: работа с байтами файла, без БД и сети.
"""
from __future__ import annotations

import io

import pytest
from openpyxl import Workbook

from app.modules.ingestion import analyze, mapping, parsers


def _letter_style_report() -> bytes:
    """xlsx в форме «приложение к письму», как у заказчика.

    Строка 1  — баннер «ИНФОРМАЦИЯ», объединён на все 4 столбца.
    Строка 2  — пустая (в оригинале такие есть, и они сдвигают нумерацию).
    Строка 3  — верхний этаж шапки, объединён над столбцами значений.
    Строка 4  — нижний этаж шапки.
    Строки 5+ — данные.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Приложение"

    ws["A1"] = "ИНФОРМАЦИЯ"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    ws["A3"] = "Субъект"
    ws["B3"] = "Показатели эффективности внедрения сервиса"
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=4)
    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)  # «Субъект» на два этажа

    ws["B4"] = "Записались"
    ws["C4"] = "Обратились"
    ws["D4"] = "Доля"

    for i, (name, a, b, c) in enumerate(
        [("Донецк", 7078, 16, 0.5), ("Макеевка", 3120, 9, 0.3), ("Горловка", 1540, 4, 0.2)],
        start=5,
    ):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=a)
        ws.cell(row=i, column=3, value=b)
        ws.cell(row=i, column=4, value=c)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture(scope="module")
def parsed():
    result = parsers.parse(_letter_style_report(), "xlsx")
    assert result.tables, "лист должен распознаться"
    return result.tables[0]


def test_grid_keeps_original_geometry(parsed):
    """Пустая строка внутри листа сохраняется: нумерация строк совпадает с Excel."""
    assert parsed.row_count == 7  # 4 «служебных» + 3 строки данных
    assert parsed.rows[1] == ["", "", "", ""]  # строка 2 листа — пустая
    assert parsed.rows[4][0] == "Донецк"  # строка 5 листа


def test_merges_are_recorded_not_tiled(parsed):
    """Значение объединения лежит только в левой верхней клетке, диапазон — в merges."""
    assert (0, 0, 0, 3) in parsed.merges  # баннер во всю ширину
    assert (2, 1, 2, 3) in parsed.merges  # верхний этаж шапки над значениями
    assert parsed.rows[0][0] == "ИНФОРМАЦИЯ"
    assert parsed.rows[0][1] == "", "баннер не должен размножаться по столбцам"


def test_fill_merges_expands_for_analysis(parsed):
    filled = parsers.fill_merges(parsed.rows, parsed.merges)
    assert filled[0] == ["ИНФОРМАЦИЯ"] * 4
    assert filled[2][3] == "Показатели эффективности внедрения сервиса"
    assert filled[3][0] == "Субъект", "вертикальное объединение спускается на нижний этаж"
    assert parsed.rows[0][1] == "", "исходная сетка не должна меняться"


def test_banner_excluded_from_data_rect(parsed):
    """Баннер и пустая строка не попадают в область данных."""
    rect = analyze.detect_data_rect(parsed.rows, parsed.merges)
    assert rect[0] == 2, "таблица начинается с верхнего этажа шапки (строка 3 листа)"
    assert rect[2] == parsed.row_count - 1


def test_headers_do_not_inherit_banner(parsed):
    """Главный дефект формы заказчика: имена столбцов больше не «ИНФОРМАЦИЯ»."""
    filled = parsers.fill_merges(parsed.rows, parsed.merges)
    rect = analyze.detect_data_rect(parsed.rows, parsed.merges)
    header_rows = analyze.guess_header_rows(filled, rect)
    assert header_rows == 2, "шапка занимает два этажа"

    cols = analyze.analyze_columns(filled, header_rows, rect)
    headers = [c.source_header for c in cols]
    assert not any("ИНФОРМАЦИЯ" in h for h in headers)
    assert headers[0] == "Субъект"
    assert headers[1] == "Показатели эффективности внедрения сервиса · Записались"
    assert len(set(headers)) == len(headers), "заголовки должны различаться"

    by_index = {c.column_index: c for c in cols}
    assert by_index[1].inferred_type == "number"
    assert by_index[0].inferred_type == "text"


def test_analysis_grid_cuts_letter_header(parsed):
    """Сетка разметки начинается с таблицы, а не с текста письма."""
    rect = analyze.detect_data_rect(parsed.rows, parsed.merges)
    area = mapping.analysis_grid(parsed.rows, parsed.merges, rect, "columns")
    assert area[0][0] == "Субъект"
    assert len(area) == 5  # 2 этажа шапки + 3 строки данных
    assert not any("ИНФОРМАЦИЯ" in cell for row in area for cell in row)


def test_analysis_grid_transposes_for_row_oriented_report(parsed):
    """Показатели в строках: область разворачивается, столбцы становятся записями."""
    rect = analyze.detect_data_rect(parsed.rows, parsed.merges)
    area = mapping.analysis_grid(parsed.rows, parsed.merges, rect, "rows")
    assert len(area) == 4, "было 4 столбца — стало 4 строки"
    assert area[0][0] == "Субъект"
    assert area[0][2] == "Донецк", "первая строка транспонированной сетки — названия"
    assert area[1][2] == "7078"


def test_data_rows_skips_excluded(parsed):
    """Строки, снятые галочкой в конструкторе, в датасет не попадают."""
    rect = analyze.detect_data_rect(parsed.rows, parsed.merges)
    area = mapping.analysis_grid(parsed.rows, parsed.merges, rect, "columns")
    rows = mapping.data_rows(area, 2)
    assert [r[0] for r in rows] == ["Донецк", "Макеевка", "Горловка"]
    # индексы исключения — в координатах сетки разметки (шапка тоже считается)
    rows = mapping.data_rows(area, 2, skip_rows=[3])
    assert [r[0] for r in rows] == ["Донецк", "Горловка"]


def test_column_numbering_row_is_header_not_data():
    """Строка нумерации граф «1 2 3» — обязательный элемент госформ.

    Раньше `guess_header_rows` останавливался на ней (в ней же числа!), и она
    уезжала в данные: первая строка датасета выглядела как «2, 3» вместо
    реальных значений. В имя показателя номер графы попадать тоже не должен.
    """
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Субъект"
    ws["B1"] = "Записались"
    ws["C1"] = "Обратились"
    ws["A2"], ws["B2"], ws["C2"] = 1, 2, 3          # нумерация граф
    ws["A3"], ws["B3"], ws["C3"] = "Донецк", 7078, 16
    ws["A4"], ws["B4"], ws["C4"] = "Макеевка", 3120, 9
    buf = io.BytesIO()
    wb.save(buf)
    tbl = parsers.parse(buf.getvalue(), "xlsx").tables[0]

    rect = analyze.detect_data_rect(tbl.rows, tbl.merges)
    filled = parsers.fill_merges(tbl.rows, tbl.merges)
    header_rows = analyze.guess_header_rows(filled, rect)
    assert header_rows == 2, "заголовок + строка нумерации граф"

    cols = analyze.analyze_columns(filled, header_rows, rect)
    assert [c.source_header for c in cols] == ["Субъект", "Записались", "Обратились"]

    area = mapping.analysis_grid(tbl.rows, tbl.merges, rect, "columns")
    assert [r[0] for r in mapping.data_rows(area, header_rows)] == ["Донецк", "Макеевка"]


def test_numbering_row_needs_consecutive_numbers():
    """Обычная строка данных с числами нумерацией граф не считается."""
    assert analyze.is_numbering_row(["1", "2", "3", "4"])
    assert not analyze.is_numbering_row(["1", "2", "5"])       # пропуск
    assert not analyze.is_numbering_row(["7078", "16", "41"])  # значения
    assert not analyze.is_numbering_row(["1", "2"])            # слишком коротко


def test_footer_rows_have_no_numbers(parsed):
    """Подвал документа (согласующие, исполнитель) распознаётся как строки без чисел.

    Именно по этому признаку конструктор предлагает снять их одной кнопкой:
    в реальном отчёте под таблицей идут ФИО и телефоны, и на дашборде от такой
    строки остаётся только мусорная категория с нулём.
    """
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Субъект"
    ws["B1"] = "Значение"
    ws["A2"] = "Донецк"
    ws["B2"] = 7078
    ws["A3"] = "Согласовано:"
    ws["A4"] = "Заместитель директора           И.И. Иванов"
    buf = io.BytesIO()
    wb.save(buf)
    tbl = parsers.parse(buf.getvalue(), "xlsx").tables[0]

    rect = analyze.detect_data_rect(tbl.rows, tbl.merges)
    area = mapping.analysis_grid(tbl.rows, tbl.merges, rect, "columns")
    cols = analyze.analyze_columns(area, 1)
    numeric = [c.column_index for c in cols if c.inferred_type == "number"]
    assert numeric, "столбец значений должен определиться как числовой"

    without_numbers = [
        i for i, row in mapping.data_row_items(area, 1)
        if not any(analyze.parse_number(row[c]) is not None for c in numeric if c < len(row))
    ]
    labels = [area[i][0] for i in without_numbers]
    assert labels == ["Согласовано:", "Заместитель директора           И.И. Иванов"]


def test_type_ignores_excluded_rows():
    """Тип столбца — по строкам, которые реально уедут в датасет.

    В бланке заказчика ФИО подписанта («Д.В.Регеда») стоит в столбце значений
    ниже таблицы. Если считать тип по всем строкам, столбец становится
    текстовым, число не попадает в value_number — и показатель на дашборде
    не считается вовсе. Исключение строки должно возвращать столбцу тип числа.
    """
    area = [
        ["Субъект", "За неделю"],   # шапка
        ["Донецк", "33748"],        # данные
        ["Заместитель руководителя", "Д.В.Регеда"],  # подпись
    ]
    both = analyze.analyze_columns(area, 1)
    assert both[1].inferred_type == "text", "с подписью столбец выглядит текстовым"

    kept = mapping.data_row_items(area, 1, skip_rows=[2])
    only_data = analyze.analyze_columns(area[:1] + [row for _i, row in kept], 1)
    assert only_data[1].inferred_type == "number"


def test_field_match_key_is_the_full_name():
    """Одна и та же форма за разные даты обязана давать ОДИНАКОВЫЕ коды полей.

    Коды обрезаны 60 символами, и у граф «Количество обращений … нарастающим
    итогом» / «… за отчётную неделю» первые 60 символов совпадают. Пока
    сопоставление с уже заведёнными полями шло по коду, второй выпуск получал
    новые коды (…_3_2), и динамика по периодам разваливалась: соседние даты
    становились разными показателями.
    """
    a = "Количество обращений за результатом оказания услуг в МФЦ · Факт · нарастающим итогом"
    b = "Количество обращений за результатом оказания услуг в МФЦ · Факт · за отчётную неделю"
    assert analyze.slug(a) == analyze.slug(b), "коды после обрезки совпадают — на этом и ломалось"
    assert mapping._norm_name(a) != mapping._norm_name(b), "полные имена различаются"
    assert mapping._norm_name("  Количество   обращений ") == mapping._norm_name("количество обращений")


def test_counter_column_is_not_a_row_label():
    """«№ п/п» — счётчик строк бланка, подписью на дашборде быть не может.

    В настоящей форме заказчика этот столбец оказывается ещё и первым
    ТЕКСТОВЫМ (в незаполненных строках там «…», сноски «*», подписи), поэтому
    наивный выбор «первый текстовый столбец» назначал подписями номера строк
    вместо названий субъектов.
    """
    assert analyze.is_counter_column("№ п/п")
    assert analyze.is_counter_column("№")
    assert analyze.is_counter_column("N п/п")
    assert analyze.is_counter_column("Порядковый номер")
    assert not analyze.is_counter_column("Субъект Российской Федерации")
    assert not analyze.is_counter_column("Количество обращений")


def test_header_newlines_are_collapsed():
    """Заголовок переносят внутри ячейки — в имени показателя переносов быть не должно."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Субъект"
    ws["B1"] = "Количество отправленных уведомлений\nо готовности результатов\nв МФЦ"
    ws["A2"], ws["B2"] = "Донецк", 7078
    buf = io.BytesIO()
    wb.save(buf)
    tbl = parsers.parse(buf.getvalue(), "xlsx").tables[0]

    cols = analyze.analyze_columns(parsers.fill_merges(tbl.rows, tbl.merges), 1)
    assert cols[1].source_header == "Количество отправленных уведомлений о готовности результатов в МФЦ"


def test_short_names_drop_common_header_prefix():
    """Имя показателя не должно начинаться с общего для всех заголовка таблицы.

    Иначе в форме разметки все показатели выглядят одинаково: различие уезжает
    за правый край поля ввода.
    """
    full = [
        "Субъект",
        "Показатели эффективности · Количество пользователей · за неделю",
        "Показатели эффективности · Количество пользователей · нарастающим итогом",
        "Показатели эффективности · Доля обращений",
    ]
    assert analyze.short_names(full) == [
        "Субъект",
        "Количество пользователей · за неделю",
        "Количество пользователей · нарастающим итогом",
        "Доля обращений",
    ]


def test_short_names_keep_last_level():
    """Общий префикс режем не до пустоты: последний уровень остаётся всегда."""
    assert analyze.short_names(["Итого · 2026", "Итого · 2025"]) == ["2026", "2025"]
    assert analyze.short_names(["Итого", "Итого"]) == ["Итого", "Итого"]


def test_slug_is_short_enough_to_type_in_a_formula():
    """Код поля пишут руками в формулах метрик — режем по границе слова."""
    long = "Количество пользователей, записавшихся на посещение МФЦ (в МАХ) · за отчётную неделю"
    code = analyze.slug(long)
    assert len(code) <= analyze.MAX_CODE_LEN
    assert not code.endswith("_")
    assert code.startswith("kolichestvo_polzovateley")
    # обрезка не должна ломать различимость — этим занимается dedupe_codes
    a = analyze.slug(long + " · нарастающим итогом")
    assert analyze.dedupe_codes([code, a]) == [code, code + "_2"]


def test_dedupe_codes_keeps_release_valid():
    """Одинаковые заголовки дают разные коды — иначе выпуск падает на unique."""
    codes = analyze.dedupe_codes(["informaciya"] * 3 + ["subekt"])
    assert codes == ["informaciya", "informaciya_2", "informaciya_3", "subekt"]
    assert len(set(codes)) == len(codes)
