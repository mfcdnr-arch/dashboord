"""Загрузка ОДНОГО еженедельного файла «ДНР_статистика_на_ДД.ММ.ГГГГ.xlsx» во
все 12 ведомств раздела «Статистика услуг ДНР» (multi-release модель дат).

Файл заказчика по каждому ведомству несёт ДВЕ отчётные даты одновременно
(«…с 01.01.2026 по <прошлая дата>» и «…по <текущая дата>» в каждом блоке
столбцов) — это устройство самой формы, не наша прихоть. Скрипт извлекает из
каждого листа-ведомства ОБЕ точки и заводит по НЕДОСТАЮЩЕЙ из них новый
`dataset_release` (код `<dept>_offices`, `reporting_period_start` = дата).
Идемпотентно: если release с этой (code, period) уже существует и активен —
пропускается, повторный запуск на том же файле ничего не задвоит. Именно
поэтому один и тот же файл можно (и нужно) прогонять и как «текущий», и
как источник данных за предыдущую дату — типичный сценарий, когда следующий
недельный файл приходит раньше, чем успели прогнать этот скрипт для прошлого.

Структура листа (проверена программно на файлах 12.08 и 19.08.2026, едина
для всех 12 ведомств): строка 3 — шапка, блок из 9 столбцов на услугу
(приоритет / оказывается / принято-прошлое / принято-текущее / прирост /
выдано-прошлое / выдано-текущее / прирост / комментарии), первый блок
начинается с 4-го столбца; строки 6.. — отделения (row_label = столбец
«МФЦ (адрес...)», НЕ «Субъект» — субъект один на все строки); строка 5 —
ИТОГО, пропускается.

Использование (внутри контейнера api, где есть openpyxl+asyncpg):
    docker cp <файл.xlsx> dashbord_api:/tmp/dnr_week.xlsx
    docker cp backend/app/modules/dnr_stats/departments.py \
        dashbord_api:/app/app/modules/dnr_stats/departments.py  # если каталог менялся
    docker exec -i dashbord_api python3 - < tools/dnr_stats/load_weekly_file.py

Если в файле появилось НОВОЕ ведомство или изменилось число услуг у
существующего — сначала дополнить `DEPARTMENTS` в
`backend/app/modules/dnr_stats/departments.py` (услуги достаются построчно из
строки 2 листа, см. `_parse_sheet`), иначе сработает `assert` и скрипт
остановится, ничего не испортив.
"""
import asyncio
import os
import re
import sys
from datetime import date

import asyncpg

# --- Править перед каждым новым файлом ---
OBJECT_ID = "c98460cf-b5ee-40af-9024-963d0b424a65"  # «ДНР — статистика услуг в МФЦ»
SOURCE_FILE = "/tmp/dnr_week.xlsx"
# ------------------------------------------

sys.path.insert(0, "/app")
from app.modules.dnr_stats.departments import DEPARTMENTS, field  # noqa: E402

SHEET_BY_CODE = {
    "mvd": "МВД", "fns": "ФНС", "rosreestr": "Росреестр", "socfond": "Соц.фонд",
    "mincifry": "Минцифры", "zags": "ЗАГС", "minjust": "Минюст", "tfoms": "ТФОМС",
    "fssp": "ФССП", "rosim": "Росимущество", "kbki": "КБКИ", "minoborony": "Минобороны",
}

RU_DATE_RE = re.compile(r"по (\d{2})\.(\d{2})\.(\d{4})")
_WS_RE = re.compile(r"\s+")


def _clean_addr(text: str) -> str:
    """Схлопывает переносы строк/повторные пробелы — как `ingestion.analyze._clean`.
    Без этого один и тот же офис в разных листах то и дело оказывался бы под
    РАЗНЫМИ ключами (лишний пробел в конце ячейки одного из ведомств), и
    список отделений/«Обзор» считали бы больше офисов, чем есть на самом деле
    (реальный случай: 71 вместо 63 — найдено и исправлено 27.08.2026)."""
    return _WS_RE.sub(" ", text).strip()


def _short_city(label: str) -> str:
    for marker in ("г. ", "г."):
        i = label.find(marker)
        if i != -1:
            return label[i + len(marker):].split(",")[0].strip()
    return label[:20]


def _parse_sheet(ws):
    """-> (n_blocks, date_prev, date_cur, offices{addr: {city, blocks:[...]}})."""
    row3 = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
    row2 = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    starts = [c + 1 for c, v in enumerate(row3) if v and "Приоритетная" in str(v)]
    names = [re.sub(r"^\d+\.\s*", "", str(v).strip()) for v in row2 if v and re.match(r"^\d+\.", str(v).strip())]
    assert len(starts) == len(names), f"block/name count mismatch: {len(starts)} vs {len(names)}"
    n_blocks = len(starts)

    m_prev = RU_DATE_RE.search(str(ws.cell(3, starts[0] + 2).value or ""))
    m_cur = RU_DATE_RE.search(str(ws.cell(3, starts[0] + 3).value or ""))
    date_prev = date(int(m_prev.group(3)), int(m_prev.group(2)), int(m_prev.group(1)))
    date_cur = date(int(m_cur.group(3)), int(m_cur.group(2)), int(m_cur.group(1)))

    offices = {}
    for r in range(6, ws.max_row + 1):
        raw_addr = ws.cell(r, 3).value
        if not raw_addr:
            continue
        addr = _clean_addr(str(raw_addr))
        blocks = []
        for start in starts:
            okaz = ws.cell(r, start + 1).value
            if hasattr(okaz, "date") and callable(getattr(okaz, "date", None)):
                okaz = okaz.date().isoformat()
            elif hasattr(okaz, "isoformat"):
                okaz = okaz.isoformat()
            elif okaz is not None:
                okaz = str(okaz)
            blocks.append({
                "prioritet": ws.cell(r, start).value,
                "okazyvaetsya": okaz,
                "prinyato_prev": ws.cell(r, start + 2).value,
                "prinyato_cur": ws.cell(r, start + 3).value,
                "vydano_prev": ws.cell(r, start + 5).value,
                "vydano_cur": ws.cell(r, start + 6).value,
                "kommentarii": ws.cell(r, start + 8).value,
            })
        offices[addr] = {"city": _short_city(addr), "blocks": blocks}
    return n_blocks, date_prev, date_cur, offices


async def _existing_release(conn, org_id, code, period):
    return await conn.fetchval(
        "select id from dataset_releases where organization_id=$1 and code=$2 "
        "and reporting_period_start=$3 and status <> 'superseded'", org_id, code, period)


async def _ensure_canonical_fields(conn, dept_code, n_blocks, admin_id):
    rows = []
    for i in range(1, n_blocks + 1):
        rows += [
            (field(dept_code, i, "prinyato"), f"Услуга {i}: Принято"),
            (field(dept_code, i, "vydano"), f"Услуга {i}: Выдано"),
            (field(dept_code, i, "prioritet"), f"Услуга {i}: Приоритетная услуга"),
            (field(dept_code, i, "okazyvaetsya"), f"Услуга {i}: Услуга оказывается"),
            (field(dept_code, i, "kommentarii"), f"Услуга {i}: Комментарии по офисам"),
        ]
    for code, name in rows:
        await conn.execute(
            "insert into canonical_fields(object_id, code, name, data_type, created_by) "
            "values($1::uuid,$2,$3,'text',$4) on conflict (object_id, code) do nothing",
            OBJECT_ID, code, name, admin_id)
    await conn.execute(
        "insert into canonical_fields(object_id, code, name, data_type, created_by) "
        "values($1::uuid,'gorod','Город','text',$2) on conflict (object_id, code) do nothing",
        OBJECT_ID, admin_id)


def _num(v):
    return float(v) if isinstance(v, (int, float)) else None


async def _build_release(conn, org_id, dept_code, dataset_code, name_suffix, period,
                          admin_id, offices, which):
    """which: 'prev' или 'cur' — какую из двух колонок файла брать за значения."""
    if await _existing_release(conn, org_id, dataset_code, period):
        return 0

    rel = await conn.fetchval(
        "insert into dataset_releases(organization_id,code,name,status,reporting_period_start,"
        "created_by,object_id) values($1,$2,$3,'validated',$4,$5,$6::uuid) returning id",
        org_id, dataset_code, f"{name_suffix} — {period.isoformat()}", period, admin_id, OBJECT_ID)

    numbers, texts = [], []
    for i, (addr, off) in enumerate(offices.items()):
        texts.append((rel, i, addr, "gorod", off["city"]))
        for svc_i, blk in enumerate(off["blocks"], start=1):
            if (pn := _num(blk[f"prinyato_{which}"])) is not None:
                numbers.append((rel, i, addr, field(dept_code, svc_i, "prinyato"), pn))
            if (vn := _num(blk[f"vydano_{which}"])) is not None:
                numbers.append((rel, i, addr, field(dept_code, svc_i, "vydano"), vn))
            if blk["prioritet"] is not None:
                texts.append((rel, i, addr, field(dept_code, svc_i, "prioritet"), str(blk["prioritet"])))
            if blk["okazyvaetsya"] is not None:
                texts.append((rel, i, addr, field(dept_code, svc_i, "okazyvaetsya"), str(blk["okazyvaetsya"])))
            if blk["kommentarii"]:
                texts.append((rel, i, addr, field(dept_code, svc_i, "kommentarii"), str(blk["kommentarii"])))

    if numbers:
        await conn.executemany(
            "insert into dataset_values(dataset_release_id,row_index,row_label,canonical_field_code,value_number) "
            "values($1,$2,$3,$4,$5)", numbers)
    if texts:
        await conn.executemany(
            "insert into dataset_values(dataset_release_id,row_index,row_label,canonical_field_code,value_text) "
            "values($1,$2,$3,$4,$5)", texts)
    return len(numbers) + len(texts)


async def main():
    import openpyxl

    conn = await asyncpg.connect(
        host=os.environ["POSTGRES_HOST"], port=int(os.environ["POSTGRES_PORT"]),
        user=os.environ["POSTGRES_USER"], password=os.environ["POSTGRES_PASSWORD"],
        database=os.environ["POSTGRES_DB"])
    org_id = await conn.fetchval("select organization_id from objects where id=$1::uuid", OBJECT_ID)
    admin_id = await conn.fetchval("select id from users where login='admin'")
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)

    for code, meta in DEPARTMENTS.items():
        sheet_name = SHEET_BY_CODE.get(code)
        if sheet_name is None or sheet_name not in wb.sheetnames:
            print(f"{code:12s} — лист не найден в файле, пропущено")
            continue
        n_blocks, date_prev, date_cur, offices = _parse_sheet(wb[sheet_name])
        assert n_blocks == len(meta["services"]), (
            f"{code}: в файле {n_blocks} услуг, в каталоге {len(meta['services'])} — "
            "обновите DEPARTMENTS в departments.py, прежде чем грузить дальше")

        await _ensure_canonical_fields(conn, code, n_blocks, admin_id)
        n1 = await _build_release(conn, org_id, code, meta["dataset_code"], meta["name"],
                                   date_prev, admin_id, offices, "prev")
        n2 = await _build_release(conn, org_id, code, meta["dataset_code"], meta["name"],
                                   date_cur, admin_id, offices, "cur")
        print(f"{code:12s} услуг={n_blocks:2d}  {date_prev} новых_знач={n1:6d}  |  {date_cur} новых_знач={n2:6d}")

    await conn.close()
    print("\nГотово. 0 новых значений у обеих дат = release'ы уже существовали (идемпотентно).")


asyncio.run(main())
